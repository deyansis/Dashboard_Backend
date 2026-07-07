from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    PageBreak,
    Table,
    TableStyle
)

from reportlab.lib.styles import (
    getSampleStyleSheet
)

from reportlab.lib.enums import (
    TA_CENTER
)
from reportlab.lib import colors
from reportlab.lib.units import cm

from openpyxl import Workbook

from datetime import datetime

import os

# ===============================
# ESTILOS
# ===============================

estilos = getSampleStyleSheet()

titulo = estilos["Title"]
titulo.alignment = TA_CENTER

subtitulo = estilos["Heading2"]

normal = estilos["BodyText"]

normal.spaceAfter = 8

# ===============================
# SENTIMIENTO PREDOMINANTE
# ===============================

def obtener_sentimiento_predominante(
    positivos,
    negativos,
    neutrales
):

    valores = {

        "Positivo": positivos,

        "Negativo": negativos,

        "Neutral": neutrales

    }

    return max(
        valores,
        key=valores.get
    )

# ===============================
# PORTADA
# ===============================

def agregar_portada(
    contenido,
    nombre,
    total,
    fecha_inicio,
    fecha_fin
):

    contenido.append(

        Paragraph(
            "SENTIGOB",
            titulo
        )

    )

    contenido.append(
        Spacer(1,20)
    )

    contenido.append(

        Paragraph(
            "Sistema Inteligente para el Análisis de Sentimientos",
            subtitulo
        )

    )

    contenido.append(
        Spacer(1,12)
    )

    contenido.append(

        Paragraph(
            "Municipalidad Distrital de Aucallama",
            normal
        )

    )

    contenido.append(
        Spacer(1,25)
    )

    contenido.append(

        Paragraph(
            f"<b>{nombre}</b>",
            subtitulo
        )

    )

    contenido.append(
        Spacer(1,15)
    )

    contenido.append(

    Paragraph(
        f"Período analizado: {fecha_inicio} al {fecha_fin}",
        normal
    )

)

    contenido.append(
    Spacer(1,8)
)

    contenido.append(

        Paragraph(
            f"Total de comentarios analizados: {total}",
            normal
        )

    )

    contenido.append(
        PageBreak()
    )
# ===============================
# COMENTARIOS
# ===============================

def agregar_comentarios(
    contenido,
    comentarios
):

    contenido.append(
        Paragraph(
            "COMENTARIOS ANALIZADOS",
            subtitulo
        )
    )

    contenido.append(
        Spacer(1, 12)
    )

    for i, comentario in enumerate(comentarios[:20], start=1):

        contenido.append(
            Paragraph(
                f"<b>Comentario #{i}</b>",
                subtitulo
            )
        )

        texto = f"""
        <b>Sentimiento:</b> {comentario['sentimiento'].capitalize()}<br/>
        <b>Prioridad:</b> {comentario['prioridad'].capitalize()}<br/><br/>
        <b>Comentario:</b><br/>
        {comentario['comentario']}
        """

        contenido.append(
            Paragraph(
                texto,
                normal
            )
        )

        contenido.append(
            Spacer(1, 15)
        )
# ===============================
# RESUMEN GENERAL
# ===============================

def agregar_resumen(
    contenido,
    total,
    positivos,
    negativos,
    neutrales
):

    contenido.append(

        Paragraph(
            "RESUMEN GENERAL",
            subtitulo
        )

    )

    contenido.append(
        Spacer(1,15)
    )

    datos = [

        ["Indicador","Cantidad"],

        ["Total de comentarios",str(total)],

        ["Comentarios positivos",str(positivos)],

        ["Comentarios negativos",str(negativos)],

        ["Comentarios neutrales",str(neutrales)]

    ]

    tabla = Table(
        datos,
        colWidths=[10*cm,4*cm]
    )

    tabla.setStyle(
        TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),1,colors.grey),
            ("BACKGROUND",(0,1),(-1,-1),colors.whitesmoke),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("BOTTOMPADDING",(0,0),(-1,0),10),
            ("TOPPADDING",(0,1),(-1,-1),8),
            ("BOTTOMPADDING",(0,1),(-1,-1),8),
            ("ALIGN",(1,1),(-1,-1),"CENTER"),
        ])
    )

    contenido.append(tabla)

    contenido.append(
        Spacer(1,20)
    )

    # sentimiento predominante

    sentimiento = obtener_sentimiento_predominante(
        positivos,
        negativos,
        neutrales
    )

    contenido.append(
        Spacer(1,15)
    )

    contenido.append(

        Paragraph(

            f"<b>Sentimiento predominante:</b> {sentimiento}",

            normal

        )

    )

# ===============================
# CONCLUSIONES
# ===============================

def agregar_conclusiones(
    contenido,
    total,
    positivos,
    negativos,
    neutrales
):

    contenido.append(

        Paragraph(
            "CONCLUSIONES",
            subtitulo
        )

    )

    contenido.append(
        Spacer(1, 12)
    )

    if positivos >= negativos and positivos >= neutrales:

        conclusion = (
            "El análisis de sentimientos evidencia una percepción "
            "mayoritariamente positiva por parte de la ciudadanía hacia "
            "la gestión municipal."
        )

    elif negativos >= positivos and negativos >= neutrales:

        conclusion = (
            "El análisis evidencia un predominio de comentarios negativos, "
            "lo que refleja un nivel importante de insatisfacción ciudadana "
            "y requiere atención prioritaria."
        )

    else:

        conclusion = (
            "La mayoría de comentarios presentan un sentimiento neutral, "
            "lo que indica opiniones informativas o sin una valoración "
            "marcadamente positiva o negativa."
        )

    contenido.append(

        Paragraph(
            conclusion,
            normal
        )

    )

    contenido.append(
        Spacer(1,15)
    )

# ===============================
# RECOMENDACIONES
# ===============================

def agregar_recomendaciones(
    contenido,
    positivos,
    negativos,
    neutrales
):

    contenido.append(

        Paragraph(
            "RECOMENDACIONES",
            subtitulo
        )

    )

    contenido.append(
        Spacer(1,12)
    )

    recomendaciones = []

    if negativos > positivos:

        recomendaciones.append(
            "• Fortalecer las acciones de atención a las principales quejas ciudadanas."
        )

        recomendaciones.append(
            "• Priorizar el seguimiento de los comentarios clasificados con prioridad alta."
        )

        recomendaciones.append(
            "• Incrementar la comunicación con la ciudadanía para mejorar la percepción institucional."
        )

    elif positivos > negativos:

        recomendaciones.append(
            "• Mantener las estrategias actuales que generan una percepción favorable."
        )

        recomendaciones.append(
            "• Continuar monitoreando las redes sociales para detectar cambios de tendencia."
        )

        recomendaciones.append(
            "• Reforzar las buenas prácticas identificadas durante el análisis."
        )

    else:

        recomendaciones.append(
            "• Continuar monitoreando la opinión ciudadana de forma periódica."
        )

        recomendaciones.append(
            "• Analizar con mayor profundidad los temas más frecuentes."
        )

    for item in recomendaciones:

        contenido.append(

            Paragraph(
                item,
                normal
            )

        )

    contenido.append(
        Spacer(1,20)
    )

# ===============================
# REPORTE COMPLETO
# ===============================

def generar_pdf_completo(
    nombre,
    comentarios,
    total,
    positivos,
    negativos,
    neutrales,
    fecha_inicio,
    fecha_fin
):

    carpeta = "reports"

    os.makedirs(
        carpeta,
        exist_ok=True
    )

    ruta = f"{carpeta}/{nombre}.pdf"

    pdf = SimpleDocTemplate(
        ruta
    )

    contenido = []

    # ==========================
    # PORTADA
    # ==========================

    agregar_portada(
    contenido,
    nombre,
    total,
    fecha_inicio,
    fecha_fin
)

    # ==========================
    # RESUMEN
    # ==========================

    agregar_resumen(
        contenido,
        total,
        positivos,
        negativos,
        neutrales
    )

    contenido.append(
        Spacer(1,20)
    )

    # ==========================
    # COMENTARIOS
    # ==========================

    agregar_comentarios(
        contenido,
        comentarios
)

    contenido.append(
    PageBreak()
)

    # ==========================
    # CONCLUSIONES
    # ==========================

    agregar_conclusiones(
        contenido,
        total,
        positivos,
        negativos,
        neutrales
    )

    # ==========================
    # RECOMENDACIONES
    # ==========================

    agregar_recomendaciones(
        contenido,
        positivos,
        negativos,
        neutrales
    )

    pdf.build(
        contenido
    )

    return ruta

# ===============================
# REPORTE EJECUTIVO
# ===============================

def generar_pdf_ejecutivo(
    nombre,
    comentarios,
    total,
    positivos,
    negativos,
    neutrales,
    fecha_inicio,
    fecha_fin
):

    carpeta = "reports"

    os.makedirs(
        carpeta,
        exist_ok=True
    )

    ruta = f"{carpeta}/{nombre}.pdf"

    pdf = SimpleDocTemplate(
        ruta
    )

    contenido = []

    agregar_portada(
        contenido,
        nombre,
        total,
        fecha_inicio,
        fecha_fin
    )

    agregar_resumen(
        contenido,
        total,
        positivos,
        negativos,
        neutrales
    )

    agregar_conclusiones(
        contenido,
        total,
        positivos,
        negativos,
        neutrales
    )

    agregar_recomendaciones(
        contenido,
        positivos,
        negativos,
        neutrales
    )

    pdf.build(
        contenido
    )

    return ruta


# ===============================
# EXCEL
# ===============================

def generar_excel(
    nombre,
    comentarios,
    total,
    positivos,
    negativos,
    neutrales,
    fecha_inicio,
    fecha_fin
):

    carpeta = "reports"

    os.makedirs(
        carpeta,
        exist_ok=True
    )

    ruta = f"{carpeta}/{nombre}.xlsx"

    wb = Workbook()

    # ==========================
    # HOJA 1 - RESUMEN
    # ==========================

    ws = wb.active

    ws.title = "Resumen"

    ws["A1"] = "SENTIGOB"

    ws["A3"] = "Nombre del reporte"
    ws["B3"] = nombre

    ws["A4"] = "Periodo analizado"
    ws["B4"] = f"{fecha_inicio} al {fecha_fin}"

    ws["A5"] = "Fecha de generación"
    ws["B5"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws["A7"] = "Total comentarios"
    ws["B7"] = total

    ws["A8"] = "Positivos"
    ws["B8"] = positivos

    ws["A9"] = "Negativos"
    ws["B9"] = negativos

    ws["A10"] = "Neutrales"
    ws["B10"] = neutrales

    # ==========================
    # HOJA 2 - COMENTARIOS
    # ==========================

    ws2 = wb.create_sheet(
        "Comentarios"
    )

    ws2.append([
        "Fecha",
        "Sentimiento",
        "Prioridad",
        "Tema",
        "Comentario"
    ])

    for comentario in comentarios:

        ws2.append([

            comentario.get(
                "fecha_registro"
            ),

            comentario.get(
                "sentimiento"
            ),

            comentario.get(
                "prioridad"
            ),

            comentario.get(
                "tema"
            ),

            comentario.get(
                "comentario"
            )

        ])

    wb.save(
        ruta
    )

    return ruta